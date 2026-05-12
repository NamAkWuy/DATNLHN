"""
Dịch vụ xuất báo cáo: Excel (openpyxl) và PDF (reportlab).
"""
import io
import os
from datetime import date, datetime
from typing import Optional


_PDF_FONT_REGULAR = "Helvetica"
_PDF_FONT_BOLD = "Helvetica-Bold"
_PDF_FONTS_REGISTERED = False


def _register_pdf_fonts() -> None:
    """Đăng ký font TTF Unicode để PDF hiển thị được dấu tiếng Việt.

    Thử lần lượt các font hệ thống phổ biến trên Windows / Linux. Nếu không tìm
    thấy thì lặng lẽ fall back về Helvetica (Type 1 sẵn có, không hỗ trợ tiếng Việt).
    """
    global _PDF_FONT_REGULAR, _PDF_FONT_BOLD, _PDF_FONTS_REGISTERED
    if _PDF_FONTS_REGISTERED:
        return

    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont

    candidates = [
        ("VN", "C:/Windows/Fonts/arial.ttf", "C:/Windows/Fonts/arialbd.ttf"),
        ("VN", "C:/Windows/Fonts/segoeui.ttf", "C:/Windows/Fonts/segoeuib.ttf"),
        ("VN", "C:/Windows/Fonts/tahoma.ttf", "C:/Windows/Fonts/tahomabd.ttf"),
        ("VN", "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
         "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf"),
        ("VN", "/Library/Fonts/Arial.ttf", "/Library/Fonts/Arial Bold.ttf"),
    ]

    for name, regular_path, bold_path in candidates:
        if os.path.exists(regular_path):
            try:
                pdfmetrics.registerFont(TTFont(name, regular_path))
                bold_name = f"{name}-Bold"
                if os.path.exists(bold_path):
                    pdfmetrics.registerFont(TTFont(bold_name, bold_path))
                else:
                    bold_name = name
                _PDF_FONT_REGULAR = name
                _PDF_FONT_BOLD = bold_name
                break
            except Exception:
                continue

    _PDF_FONTS_REGISTERED = True


# ---------------------------------------------------------------------------
# Xuất file Excel
# ---------------------------------------------------------------------------

def generate_attendance_excel(
    month: int,
    year: int,
    records: list[dict],
) -> bytes:
    """
    Tạo file Excel chứa dữ liệu báo cáo chấm công.
    Trả về bytes thô của file .xlsx.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = f"Cham cong {month:02d}-{year}"

    # ---- Tiêu đề ----
    title_font = Font(name="Arial", bold=True, size=14)
    header_font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    ws.merge_cells("A1:J1")
    title_cell = ws["A1"]
    title_cell.value = f"BÁO CÁO CHẤM CÔNG THÁNG {month:02d}/{year}"
    title_cell.font = title_font
    title_cell.alignment = center_align

    ws.merge_cells("A2:J2")
    subtitle = ws["A2"]
    subtitle.value = f"Xuất ngày: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    subtitle.alignment = center_align

    # ---- Dòng tiêu đề cột ----
    headers = [
        "STT", "Mã NV", "Họ và tên", "Phòng ban",
        "Ngày công", "Số giờ làm", "Đi muộn", "Về sớm", "Vắng", "Ngày phép"
    ]
    col_widths = [6, 10, 24, 18, 10, 12, 10, 10, 8, 10]

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = col_widths[col_idx - 1]

    # ---- Các dòng dữ liệu ----
    alt_fill = PatternFill(start_color="EFF6FF", end_color="EFF6FF", fill_type="solid")
    data_font = Font(name="Arial", size=10)

    for row_idx, record in enumerate(records, start=1):
        row_num = row_idx + 4
        fill = alt_fill if row_idx % 2 == 0 else PatternFill()
        values = [
            row_idx,
            record.get("employee_code", ""),
            record.get("full_name", ""),
            record.get("department", ""),
            record.get("total_days_worked", 0),
            record.get("total_work_hours", 0),
            record.get("late_count", 0),
            record.get("early_leave_count", 0),
            record.get("absent_count", 0),
            record.get("leave_days", 0),
        ]
        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_num, column=col_idx, value=value)
            cell.font = data_font
            cell.border = thin_border
            cell.fill = fill
            if col_idx == 1 or col_idx >= 5:
                cell.alignment = center_align

    # ---- Chiều cao dòng ----
    ws.row_dimensions[1].height = 28
    ws.row_dimensions[4].height = 20

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


# ---------------------------------------------------------------------------
# Xuất file PDF
# ---------------------------------------------------------------------------

def generate_attendance_pdf(
    month: int,
    year: int,
    records: list[dict],
) -> bytes:
    """
    Tạo báo cáo PDF bằng reportlab.
    Trả về bytes thô của file PDF.
    """
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.enums import TA_CENTER

    _register_pdf_fonts()

    output = io.BytesIO()
    doc = SimpleDocTemplate(
        output,
        pagesize=landscape(A4),
        rightMargin=1.5 * cm,
        leftMargin=1.5 * cm,
        topMargin=1.5 * cm,
        bottomMargin=1.5 * cm,
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "CustomTitle",
        parent=styles["Title"],
        fontName=_PDF_FONT_BOLD,
        fontSize=16,
        spaceAfter=6,
        alignment=TA_CENTER,
    )
    subtitle_style = ParagraphStyle(
        "CustomSubtitle",
        parent=styles["Normal"],
        fontName=_PDF_FONT_REGULAR,
        fontSize=10,
        spaceAfter=12,
        alignment=TA_CENTER,
    )

    elements = []

    # Tiêu đề
    elements.append(Paragraph(f"BÁO CÁO CHẤM CÔNG THÁNG {month:02d}/{year}", title_style))
    elements.append(Paragraph(
        f"Xuất ngày: {datetime.now().strftime('%d/%m/%Y %H:%M')}",
        subtitle_style,
    ))
    elements.append(Spacer(1, 0.3 * cm))

    # Tiêu đề các cột bảng
    col_headers = [
        "STT", "Mã NV", "Họ và tên", "Phòng ban",
        "Ngày công", "Số giờ làm", "Đi muộn", "Về sớm", "Vắng", "Ngày phép"
    ]
    col_widths = [1.2 * cm, 2.0 * cm, 5.0 * cm, 3.5 * cm,
                  2.0 * cm, 2.2 * cm, 2.0 * cm, 2.0 * cm, 1.8 * cm, 2.2 * cm]

    table_data = [col_headers]
    for row_idx, record in enumerate(records, start=1):
        table_data.append([
            str(row_idx),
            record.get("employee_code", ""),
            record.get("full_name", ""),
            record.get("department", "") or "",
            str(record.get("total_days_worked", 0)),
            f"{record.get('total_work_hours', 0):g}",
            str(record.get("late_count", 0)),
            str(record.get("early_leave_count", 0)),
            str(record.get("absent_count", 0)),
            str(record.get("leave_days", 0)),
        ])

    table = Table(table_data, colWidths=col_widths, repeatRows=1)
    table.setStyle(TableStyle([
        # Dòng tiêu đề
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2563EB")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), _PDF_FONT_BOLD),
        ("FONTSIZE", (0, 0), (-1, 0), 9),
        ("ALIGN", (0, 0), (-1, 0), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        # Các dòng dữ liệu
        ("FONTNAME", (0, 1), (-1, -1), _PDF_FONT_REGULAR),
        ("FONTSIZE", (0, 1), (-1, -1), 8),
        ("ALIGN", (0, 1), (1, -1), "CENTER"),
        ("ALIGN", (4, 1), (-1, -1), "CENTER"),
        # Tô màu xen kẽ các dòng
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#EFF6FF")]),
        # Đường kẻ ô
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
    ]))

    elements.append(table)

    # Dòng tổng hợp ở cuối báo cáo
    elements.append(Spacer(1, 0.5 * cm))
    total_present = sum(r.get("total_days_worked", 0) for r in records)
    total_hours = sum(r.get("total_work_hours", 0) for r in records)
    summary_text = (
        f"Tổng số nhân viên: {len(records)} | "
        f"Tổng ngày công: {total_present} | "
        f"Tổng giờ làm: {total_hours:g}"
    )
    elements.append(Paragraph(summary_text, subtitle_style))

    doc.build(elements)
    return output.getvalue()
